import os
import glob
import pandas as pd
from datetime import datetime
from plyer import notification
import tkinter as tk
from tkinter import simpledialog
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- Infraestrutura ---
DIRETORIO_BASE = r"C:\Sistemas\MotorFinanceiro\planilhas" 
RECEITA_LIQUIDA = 1578.07

def auditar_fatura_projetada():
    """Interceção de dados via UI."""
    root = tk.Tk()
    root.withdraw()
    try:
        fatura = simpledialog.askfloat(
            "Controle Mensal", 
            "Fatura do cartão de crédito (Apenas a sua parte):", 
            parent=root
        )
        return fatura if fatura is not None else 0.0
    except Exception:
        return 0.0
    finally:
        root.destroy()

def obter_ultimo_banco_de_dados():
    arquivos_excel = glob.glob(os.path.join(DIRETORIO_BASE, "*.xlsx"))
    return max(arquivos_excel, key=os.path.getmtime) if arquivos_excel else None

def inicializar_ciclo_financeiro():
    if not os.path.exists(DIRETORIO_BASE):
        os.makedirs(DIRETORIO_BASE)
        
    data_atual = datetime.now().strftime("%d-%m-%Y")
    nome_ficheiro = f"Controle_Pessoal_{data_atual}.xlsx"
    caminho_completo = os.path.join(DIRETORIO_BASE, nome_ficheiro)
    
    fatura_projetada = auditar_fatura_projetada()
    
    # Processo de extração da planilha anterior (se existir)
    planilha_anterior = obter_ultimo_banco_de_dados()
    df = None

    if planilha_anterior:
        try:
            df = pd.read_excel(planilha_anterior)
            # Zera apenas o que foi gasto, mantém o planejamento
            if "Valor Gasto (Real)" in df.columns:
                df["Valor Gasto (Real)"] = 0
                
            idx_cartao = df.index[df['Descrição'] == 'Cartão de Crédito'].tolist()
            if idx_cartao:
                df.at[idx_cartao[0], 'Valor Planejado'] = fatura_projetada
        except Exception:
            df = None 

    # Fallback: Matriz limpa e estritamente focada no seu escopo pessoal
    if df is None:
        colunas = ["Categoria", "Descrição", "Valor Planejado", "Valor Gasto (Real)"]
        dados_matriz = [
            ["Custos Fixos", "Faculdade", 527.64, 0],
            ["Custos Fixos", "Internet", 59.90, 0],
            ["Custos Fixos", "Cartão de Crédito", fatura_projetada, 0],
            ["Alimentação", "Lanches/Mercado", 0, 0],
            ["Lazer", "Spotify", 23.90, 0],
            ["Lazer", "Saídas de Final de Semana", 0, 0],
            ["Investimentos", "Câmbio / Caixinhas", 400.00, 0]
        ]
        df = pd.DataFrame(dados_matriz, columns=colunas)
    
    # 1. Salva a estrutura base com Pandas
    try:
        df.to_excel(caminho_completo, index=False, engine='openpyxl')
    except Exception as e:
        print(f"Falha de I/O: {e}")
        return

    # 2. Injeta Design e Automação diretamente no Motor do Excel (OpenPyxl)
    wb = load_workbook(caminho_completo)
    ws = wb.active
    ws.title = "Gestão Pessoal"

    # Estilo do Cabeçalho (Azul escuro, letra branca e negrito)
    cor_fundo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)
    
    for col_num in range(1, 5):
        celula = ws.cell(row=1, column=col_num)
        celula.fill = cor_fundo
        celula.font = fonte_branca
        celula.alignment = Alignment(horizontal="center")

    # Ajuste automático da largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Determina onde a tabela termina para injetar as fórmulas de Total
    ultima_linha_dados = len(df) + 1
    linha_total = ultima_linha_dados + 2
    linha_saldo = linha_total + 1

    # Injeção das Fórmulas Nativas de Soma
    ws[f"B{linha_total}"] = "TOTAL DE GASTOS:"
    ws[f"B{linha_total}"].font = Font(bold=True)
    ws[f"C{linha_total}"] = f"=SUM(C2:C{ultima_linha_dados})"
    ws[f"D{linha_total}"] = f"=SUM(D2:D{ultima_linha_dados})"

    # Injeção das Fórmulas de Saldo (Receita - Gastos)
    ws[f"B{linha_saldo}"] = "SALDO RESTANTE:"
    ws[f"B{linha_saldo}"].font = Font(bold=True)
    ws[f"C{linha_saldo}"] = f"={RECEITA_LIQUIDA} - C{linha_total}"
    ws[f"D{linha_saldo}"] = f"={RECEITA_LIQUIDA} - D{linha_total}"

    # Aplica formatação de Moeda (R$) para todas as células numéricas
    for row in range(2, linha_saldo + 1):
        ws[f"C{row}"].number_format = '"R$" #,##0.00'
        ws[f"D{row}"].number_format = '"R$" #,##0.00'

    wb.save(caminho_completo)

    notification.notify(
        title="Controle Pessoal Gerado",
        message=f"Planilha otimizada. Visão limpa ativada.",
        app_name="Motor Financeiro",
        timeout=5
    )

if __name__ == "__main__":
    inicializar_ciclo_financeiro()